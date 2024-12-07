import os
import openpyxl
import re
from tkinter import Tk, Label, Entry, Button, messagebox

# Create or open the 'submissions.xlsx' file
def create_or_open_workbook():
    if not os.path.exists("submissions.xlsx"):
        # Create new workbook and sheet
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "Submissions"
        # Write headers to the sheet
        sheet.append(["ID", "Name", "Email", "Age", "Address", "Phone"])
        wb.save("submissions.xlsx")
    else:
        wb = openpyxl.load_workbook("submissions.xlsx")
    
    return wb

# Get the next auto-incremented ID
def get_next_id(wb):
    sheet = wb["Submissions"]
    max_row = sheet.max_row
    return max_row if max_row == 1 else sheet.cell(row=max_row, column=1).value + 1

# Validate email format
def validate_email(email):
    email_regex = r'^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$'
    return re.match(email_regex, email) is not None

# Save submission to Excel file
def save_submission(name, email, age, address, phone):
    wb = create_or_open_workbook()
    sheet = wb["Submissions"]
    
    # Get the next available ID
    next_id = get_next_id(wb)

    # Add data to the Excel file
    sheet.append([next_id, name, email, age, address, phone])
    wb.save("submissions.xlsx")

# Handle form submission
def on_submit(name_entry, email_entry, age_entry, address_entry, phone_entry):
    # Retrieve data from form
    name = name_entry.get()
    email = email_entry.get()
    age = age_entry.get()
    address = address_entry.get()
    phone = phone_entry.get()

    # Validate that all fields are filled
    if not name or not email or not age or not address or not phone:
        messagebox.showerror("Error", "All fields are required.")
        return
    
    # Validate email format
    if not validate_email(email):
        messagebox.showerror("Error", "Invalid email address.")
        return
    
    # Save to Excel
    save_submission(name, email, age, address, phone)

    # Clear the fields
    name_entry.delete(0, 'end')
    email_entry.delete(0, 'end')
    age_entry.delete(0, 'end')
    address_entry.delete(0, 'end')
    phone_entry.delete(0, 'end')

    # Show success message
    messagebox.showinfo("Success", "Submission saved successfully!")

# Create the GUI (form)
def create_form():
    # Create the main window
    root = Tk()
    root.title("Data Entry Form")

    # Labels
    Label(root, text="Name:").grid(row=0, column=0, padx=10, pady=5)
    Label(root, text="Email:").grid(row=1, column=0, padx=10, pady=5)
    Label(root, text="Age:").grid(row=2, column=0, padx=10, pady=5)
    Label(root, text="Address:").grid(row=3, column=0, padx=10, pady=5)
    Label(root, text="Phone:").grid(row=4, column=0, padx=10, pady=5)

    # Entries
    name_entry = Entry(root, width=30)
    name_entry.grid(row=0, column=1, padx=10, pady=5)
    email_entry = Entry(root, width=30)
    email_entry.grid(row=1, column=1, padx=10, pady=5)
    age_entry = Entry(root, width=30)
    age_entry.grid(row=2, column=1, padx=10, pady=5)
    address_entry = Entry(root, width=30)
    address_entry.grid(row=3, column=1, padx=10, pady=5)
    phone_entry = Entry(root, width=30)
    phone_entry.grid(row=4, column=1, padx=10, pady=5)

    # Submit button
    submit_button = Button(root, text="Submit", width=20, command=lambda: on_submit(name_entry, email_entry, age_entry, address_entry, phone_entry))
    submit_button.grid(row=5, column=0, columnspan=2, pady=10)

    # Run the Tkinter event loop
    root.mainloop()

# Start the form
if __name__ == "__main__":
    create_form()
